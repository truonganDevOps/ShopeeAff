from __future__ import annotations

import csv
import json
from typing import Any

EXPORT_FIELDS = [
    "score",
    "category",
    "name",
    "price",
    "sold_count",
    "rating",
    "shop_type",
    "product_link",
]

HEADERS = {
    "score":        "Score",
    "category":     "Danh Mục",
    "name":         "Tên Sản Phẩm",
    "price":        "Giá",
    "sold_count":   "Đã Bán",
    "rating":       "Rating",
    "shop_type":    "Loại Shop",
    "product_link": "Link Sản Phẩm",
}

COLUMN_WIDTHS = {
    "score":        8,
    "category":     22,
    "name":         48,
    "price":        18,
    "sold_count":   14,
    "rating":       10,
    "shop_type":    14,
    "product_link": 60,
}


def export_json(
    products: list[dict[str, Any]],
    filename: str = "products.json",
) -> None:
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(_normalize(products), f, ensure_ascii=False, indent=2)
    print(f"  JSON -> {filename} ({len(products)} items)")


def export_csv(
    products: list[dict[str, Any]],
    filename: str = "products.csv",
) -> None:
    with open(filename, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
        writer.writerow(HEADERS)
        writer.writerows(_normalize(products))
    print(f"  CSV  -> {filename} ({len(products)} items)")


def export_excel(
    overall: list[dict[str, Any]],
    per_cat: list[dict[str, Any]],
    filename: str = "products.xlsx",
) -> None:
    from openpyxl import Workbook

    wb = Workbook()

    _write_sheet(wb.active, "Top Overall", overall)
    _write_sheet(wb.create_sheet("Per Category"), "Per Category", per_cat)

    wb.save(filename)
    print(f"  Excel-> {filename}  (overall={len(overall)}, per_cat={len(per_cat)})")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _write_sheet(ws, title: str, products: list[dict[str, Any]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.title = title

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
    center      = Alignment(horizontal="center", vertical="center")
    wrap        = Alignment(wrap_text=True, vertical="top")
    top         = Alignment(vertical="top")

    ws.append([HEADERS[f] for f in EXPORT_FIELDS])
    for col_idx, field in enumerate(EXPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[field]
    ws.row_dimensions[1].height = 22

    shop_colors = {
        "mall":      "FFF2CC",
        "preferred": "E2EFDA",
        "normal":    "FFFFFF",
    }

    for row_idx, product in enumerate(_normalize(products), start=2):
        shop_type = product.get("shop_type", "normal")
        row_fill  = PatternFill(fill_type="solid", fgColor=shop_colors.get(shop_type, "FFFFFF"))

        if row_idx % 2 == 0 and shop_type == "normal":
            row_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")

        for col_idx, field in enumerate(EXPORT_FIELDS, start=1):
            value = product.get(field, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill

            if field == "product_link" and value:
                cell.hyperlink = value
                cell.font      = Font(color="0563C1", underline="single")
                cell.alignment = wrap
            elif field == "name":
                cell.alignment = wrap
            else:
                cell.alignment = top

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _normalize(products: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [
        {field: str(product.get(field, "")) for field in EXPORT_FIELDS}
        for product in products
    ]
